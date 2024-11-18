from django.shortcuts import render, redirect
from .models import Employee_Record, Client, Department, Component, Statistics
from .forms import ReceivedModelFormSet
from .filters import EmployeeRecordFilter
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
import os

#for export excel imports
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import *
from urllib.parse import quote


# Create your views here.

@login_required
def summary_released(request):
    records = Employee_Record.objects.all()

    records_Filter = EmployeeRecordFilter(request.GET, queryset=records)
    records = records_Filter.qs
    records_count = records.count()

    if records_count > 0:
        messages.info(request, f"Found '{records_count}' record(s) in the database")
    else:
        messages.info(request, f"Record not Found in the database ")


    #this will be called if user report is filtered
    global filter_employee_record_val
    def filter_employee_record_val():
        return records    

    context = {
        'records': records,
        'records_count': records_count,
        'records_Filter': records_Filter
        }
    return render(request, 'receiving_app/summary_report.html', context)



@login_required
def add_record(request):

    # create a list to hold the first value input in the formset
    employee_id_list = []
    first_name_list = []
    last_name_list = []

    if request.method == 'POST':
        formset = ReceivedModelFormSet(request.POST)
        #check the validity of form
        if formset.is_valid():
            for form in formset:

                #assign input value to variables
                add_client = form.cleaned_data.get('client')
                add_department = form.cleaned_data.get('department')
                add_employee_id = form.cleaned_data.get('employee_id')
                add_first_name = form.cleaned_data.get('first_name')
                add_last_name = form.cleaned_data.get('last_name')


                try:
                    #assign value to foreign keys
                    client = Client.objects.get(client=add_client)
                    department = Department.objects.get(department=add_department)

                except:
                    department = Department.objects.get(department="NONE")


                #hold the first value to a list
                employee_id_list.append(add_employee_id)
                first_name_list.append(add_first_name)
                last_name_list.append(add_last_name)

                #assign to each form
                count = 1

                #assign value to fields
                received_formset = form.save(commit=False)
                received_formset.client = client            
                received_formset.employee_id = employee_id_list[0]
                received_formset.first_name = first_name_list[0]
                received_formset.last_name = last_name_list[0]
                received_formset.count = count
                received_formset.department = department

                #check if component value is null
                if not form.cleaned_data.get('component'):
                    component = Component.objects.get(component="OTHERS")
                    received_formset.component = component

                # #check if department value is null
                # if not form.cleaned_data.get('department'):
                #     department = Department.objects.get(department="NONE")
                #     received_formset.department = department
                # else:
                #     received_formset.department = department
             

                received_formset.save()

            #Count each transaction per employee
            stats = Statistics(employee=employee_id_list[0], count=1)
            stats.save()

            return redirect('submitted')

    else:
        formset = ReceivedModelFormSet(queryset=Employee_Record.objects.none())

    context = {'formset': formset}
    return render(request, 'receiving_app/add_record.html', context)


@login_required
def submitted(request):
    return render(request, 'receiving_app/submitted.html')


@login_required
def dashboard(request):
    return render(request, 'receiving_app/dashboard.html')
 

@login_required
def delete_employee(request, id):
    if request.method == 'POST':
        employee = Employee_Record.objects.get(id=id)

        employee.delete()
    return redirect('summary_received')


@login_required
def export_excel_record(request):

    #Export excel function
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="HR_REPORT.xlsx"'

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


    # Declare Workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells('A1:F1')

    first_cell = worksheet['A1']
    first_cell.value = "HR_REPORT"
    first_cell.font = Font(bold=True)
    first_cell.alignment = Alignment(horizontal="center", vertical="center")


    worksheet.title = "ACKNOWLEDGEMENT REPORT"

    # Add headers
    headers =   [
                'EMPLOYEE ID',	
                'NAME',
                'CLIENT',	
                'DEPARTMENT',
                'RECEIVED',
                'DATE',
                ]
    row_num = 2


    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="CFE2FF")
        cell.font = Font(bold=True, color="0B5ED7")
        cell.border = thin_border


    # check if it was filtered if not set default to all
    try:
        employee_records = filter_employee_record_val()
        print("With filter value")
              
    except:
        employee_records = Employee_Record.objects.all()
        print("filter no value")
        

    for employee in employee_records:
        
        #convert object fields to string)
        client = str(employee.client)
        department = str(employee.department)
        component = str(employee.component)
        date = datetime.strftime(employee.date,'%m/%d/%Y %H:%M:%S')
        name = employee.last_name + ", " + employee.first_name

        worksheet.append([
            employee.employee_id,
            name, 
            client,
            department,
            component,
            date
        ])
    
    workbook.save(response)
    return response